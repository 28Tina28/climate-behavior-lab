import csv
import os
import re
from datetime import datetime, date, timedelta
from typing import Optional
import openpyxl
from .models import ClimateRecord, BehaviorRecord, Batch

CSV_COLUMN_MAP = {
    "FORMATTED DATE_TIME": "timestamp",
    "Temperature": "temperature",
    "Wet Bulb Temp": "wet_bulb_temp",
    "Globe Temperature": "globe_temperature",
    "Relative Humidity": "relative_humidity",
    "Barometric Pressure": "barometric_pressure",
    "Altitude": "altitude",
    "Station Pressure": "station_pressure",
    "Wind Speed": "wind_speed",
    "Heat Index": "heat_index",
    "Dew Point": "dew_point",
    "Density Altitude": "density_altitude",
    "Crosswind": "crosswind",
    "Headwind": "headwind",
    "Compass Magnetic Direction": "compass_magnetic",
    "NWB Temp": "nwb_temp",
    "Compass True Direction": "compass_true",
    "Thermal Work Limit": "thermal_work_limit",
    "Wet Bulb Globe Temperature": "wbgt",
    "Wind Chill": "wind_chill",
    "Time": "timestamp",
    "Temp": "temperature",
    "Wet Bulb Temp.": "wet_bulb_temp",
    "Globe Temp": "globe_temperature",
    "Rel. Hum.": "relative_humidity",
    "Baro.": "barometric_pressure",
    "Station P.": "station_pressure",
    "Dens. Alt.": "density_altitude",
    "Mag. Dir.": "compass_magnetic",
    "NA WBGT": "nwb_temp",
    "True Dir.": "compass_true",
    "TWL": "thermal_work_limit",
}

def parse_climate_csv(filepath):
    rows = []
    meta = {}
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        lines = list(reader)
    if not lines:
        return rows, meta
    for i, row in enumerate(lines):
        if i == 0 and row:
            meta["device_name"] = row[1] if len(row) > 1 else ""
        elif i == 1 and row:
            meta["device_model"] = row[1] if len(row) > 1 else ""
        elif i == 2 and row:
            meta["serial_number"] = row[1] if len(row) > 1 else ""
    header_idx = None
    for i, row in enumerate(lines):
        if row and row[0] and ("FORMATTED DATE_TIME" in str(row[0]) or str(row[0]).strip() == "Time"):
            header_idx = i
            break
    if header_idx is None:
        return rows, meta
    headers = [h.strip() for h in lines[header_idx]]
    data_start = header_idx + 2
    for row in lines[data_start:]:
        if not row or not row[0].strip():
            continue
        record = {}
        for i, h in enumerate(headers):
            if i < len(row) and h in CSV_COLUMN_MAP:
                val = row[i].strip()
                if val and val != "--" and val != "***":
                    try:
                        record[CSV_COLUMN_MAP[h]] = float(val)
                    except ValueError:
                        record[CSV_COLUMN_MAP[h]] = val
                else:
                    record[CSV_COLUMN_MAP[h]] = None
        ts_str = row[0].strip() if row else ""
        import re
        # Keep AM/PM markers, only strip other non-standard trailing text
        if re.search(r"(?:AM|PM)\s*$", ts_str, re.IGNORECASE):
            ts_clean = ts_str
        else:
            ts_clean = re.sub(r"[^\d\-: /]+$", "", ts_str).strip()
        if ts_clean:
            try:
                record["timestamp"] = datetime.strptime(ts_clean, "%Y-%m-%d %I:%M:%S %p")
            except ValueError:
                try:
                    record["timestamp"] = datetime.strptime(ts_clean, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        record["timestamp"] = datetime.strptime(ts_clean, "%Y/%m/%d %H:%M")
                    except ValueError:
                        record["timestamp"] = None
        else:
            record["timestamp"] = None
        if record.get("timestamp"):
            rows.append(record)
    return rows, meta

from datetime import time as _dtime

def parse_observation_xlsx(filepath, date_hint=None):
    """Parse observation XLSX with dynamic column detection."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not rows_iter:
        return []

    h1 = rows_iter[0] if rows_iter else []
    h2 = rows_iter[1] if len(rows_iter) > 1 else []

    # Detect column layout from headers
    def _find(kws):
        h1l, h2l = list(h1) if isinstance(h1, tuple) else h1, list(h2) if isinstance(h2, tuple) else h2
        for i, (a, b) in enumerate(zip(h1l + [None]*20, h2l + [None]*20)):
            for kw in kws:
                if (a is not None and kw in str(a)) or (b is not None and kw in str(b)):
                    return i
        return None

    col_start = _find(["起始时间", "时间"])
    col_end = _find(["结束时间"])
    col_gender = _find(["性别"])
    col_age = _find(["年龄"])
    col_color = _find(["颜色", "衣着"])
    col_thick = _find(["厚薄"])
    col_cover = _find(["遮蔽"])
    col_env = _find(["椅子环境"])
    col_chair = _find(["椅子编号"])
    col_behav = _find(["行为"])
    col_detail = _find(["具体", "聊天", "休息", "社交"])  # try to find behavior detail column
    col_weather = _find(["天气"])
    col_adjust = _find(["主动调整"])
    col_orient = _find(["朝向"])
    col_people = _find(["人数"])

    # Fallback if detection fails: try by format
    if col_gender is None:
        ncols = max(len(r) for r in rows_iter[2:] if r and any(c is not None for c in r))
        off = 1 if ncols >= 17 and h1 and h1[0] and "序号" in str(h1[0]) else 0
        col_start, col_end = off, off + 1
        col_gender = off + 2; col_age = off + 3
        col_color = off + 4; col_thick = off + 5; col_cover = off + 6
        col_env = off + 7; col_chair = off + 8
        col_behav = off + 9; col_detail = off + 10
        col_weather = off + 11
        col_adjust = off + 13; col_orient = off + 14
        col_people = off + 15

    def _cv(idx):
        if idx is None:
            return None
        return rows_iter[row_idx][idx] if row_idx < len(rows_iter) and idx < len(rows_iter[row_idx]) else None

    def _to_dt(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, date) and not isinstance(val, datetime):
            return datetime.combine(val, datetime.min.time())
        if isinstance(val, _dtime):
         return datetime.combine(date_hint or datetime.now().date(), val)
        if isinstance(val, (int, float)):
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        return None

    records = []
    for row_idx in range(2, len(rows_iter)):
        row = rows_iter[row_idx]
        if not row or not any(c is not None for c in row):
            continue

        start = _to_dt(row[col_start] if col_start is not None and col_start < len(row) else None)
        end = _to_dt(row[col_end] if col_end is not None and col_end < len(row) else None)

        def val(idx):
            return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else None
        def nval(idx):
            try: return int(float(str(row[idx]).strip())) if idx is not None and idx < len(row) and row[idx] is not None else None
            except: return None

        records.append({
            "start_time": start, "end_time": end,
            "gender": val(col_gender), "age": val(col_age),
            "clothing_color": val(col_color), "clothing_thickness": val(col_thick), "clothing_coverage": val(col_cover),
            "chair_environment": val(col_env), "chair_number": nval(col_chair),
            "behavior_category": val(col_behav), "behavior_detail": val(col_detail),
            "weather_change_from": val(col_weather),
            "active_adjustments": val(col_adjust),
            "orientation": val(col_orient), "num_people": val(col_people),
        })
    return records

def infer_batch_meta_from_filenames(csv_name, xlsx_name):
    date_label = ""
    time_period = ""
    m = re.search(r"(\d{4})(\d{2})(\d{2})", csv_name)
    if m:
        date_label = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    for kw in ["傍晚", "中午", "下午", "上午", "早晨", "凌晨", "夜间", "晚上", "午后"]:
        if kw in csv_name or kw in xlsx_name:
            time_period = kw
            break
    return {"date_label": date_label, "time_period": time_period}

def _safe_str(val):
    if val is None: return None
    s = str(val).strip()
    return s if s else None

def _safe_int(val):
    if val is None: return None
    try: return int(float(str(val).strip()))
    except: return None
